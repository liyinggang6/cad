from openpyxl import load_workbook, Workbook
from ALLPartData import PartsData
import re
class TMJSYQ():
    def __init__(self):
        self.allpart_data=PartsData()
        self.Material=self.allpart_data.Material
        self.SMaterial=self.allpart_data.SMaterial
        self.Standard=self.allpart_data.Standard

    def find_numbers(self, text):
        # 正则表达式匹配整数或小数
        pattern = r'-?\d+\.?\d*'
        # 查找所有匹配项
        numbers = re.findall(pattern, text)
        # 将匹配到的字符串转换为浮点数
        numbers = [float(num) if '.' in num else int(num) for num in numbers]
        return numbers

    def get_index_or_none(self, lst, value):
        try:
            return lst.index(value)
        except ValueError:
            return None

    def split_by_repeated_chars(self, input_string, delimiter_chars):
        """
        将输入字符串按照delimiter_chars中的一个或多个字符及其重复进行分割。

        参数:
        input_string (str): 要分割的字符串。
        delimiter_chars (str): 包含分割符字符的字符串。

        返回:
        list: 分割后的字符串列表。
        """
        # 创建一个正则表达式，匹配delimiter_chars中的任意一个字符，以及这些字符的重复
        pattern = '|'.join(re.escape(char) + '+' for char in delimiter_chars)
        # 使用re.split进行分割
        result = re.split(pattern, input_string)
        return result  

    def get_JSYQmianban(self):
        print("读取模拟界面数据库，请稍候...")
        wb = load_workbook('database/TMJSYQv5.xlsm', data_only=True)
        ws1 = wb["编码"]
        JSYQbianma = []
        for i in range(5):
            b=[]
            for j in range(12):
                a=ws1.cell(row=j+3, column=i+2).value
                if a:
                    b.append(str(a))
            JSYQbianma.append(b)
        JSYQjiemian = []
        for i in range(12):
            a=ws1.cell(row=i+17, column=4).value
            if a:
                JSYQjiemian.append(str(a))
            else:
                JSYQjiemian.append('')

        print("JSYQbianma=\n",JSYQbianma)
        print("JSYQjiemian=\n",JSYQjiemian)
        JSYQjiemianbianma = []
        JSYQjiemianbianma.append(self.get_index_or_none(JSYQbianma[0], JSYQjiemian[0]))
        JSYQjiemianbianma.append(self.get_index_or_none(JSYQbianma[1], JSYQjiemian[1]))
        fushi = self.split_by_repeated_chars(JSYQjiemian[2], ', /，\\')
        try:
            fushi.remove('')
        except ValueError:
            pass
        JSYQjiemianbianma.append([self.get_index_or_none(JSYQbianma[2],i) for i in fushi])
        yibao = self.split_by_repeated_chars(JSYQjiemian[3], ', /，\\')
        try:
            yibao.remove('')
        except ValueError:
            pass
        JSYQjiemianbianma.append([self.get_index_or_none(JSYQbianma[3],i) for i in yibao])
        JSYQjiemianbianma.append(self.get_index_or_none(JSYQbianma[4], JSYQjiemian[4]))
        JSYQjiemianbianma.append(self.find_numbers(JSYQjiemian[5]))
        JSYQjiemianbianma.append(self.find_numbers(JSYQjiemian[6]))
        JSYQjiemianbianma.append(self.find_numbers(JSYQjiemian[7])[0])
        JSYQjiemianbianma.append(self.find_numbers(JSYQjiemian[8])[0])
        JSYQjiemianbianma.append(self.find_numbers(JSYQjiemian[9])[0])
        JSYQjiemianbianma.append(self.find_numbers(JSYQjiemian[10])[0])
        JSYQjiemianbianma.append(self.find_numbers(JSYQjiemian[11])[0])
        return JSYQjiemianbianma

    def get_material_jsyq(self,data):
        ID_Material = data.get('jiceng',None)
        ID_SMaterial =  data.get('fuceng',None)
        C =  data.get('C',None)
        Standard = None
        SStandard = None
        if ID_Material:
            for i in self.Material: #获取基层材料的类型和名称
                if i[0] == ID_Material:
                    Mtype =  i[1]  #材料类型
                    Standard = i[3]
                    Mname =  i[4]  #材料名称
                    Tem_dc_low = i[7]
                    Tem_dc_nom =  i[8]
                    Kv2 =  i[9]
        m_jsyq={}            
        if ID_SMaterial:
            for i in self.SMaterial: #获取复层材料的类型和名称
                if i[0] == ID_SMaterial:
                    SMtype =  i[1]  #复合层材料类型
                    SStandard = i[3]
                    SMname =  i[4]  #复合层材料名称
                    if '堆焊' in i[1]:
                        m_jsyq['fujia'] = '堆焊' + SMname
                    if '复合层' in i[1]:
                        m_jsyq['fujia'] = '复层材料{}应符合{}的规定'.format(SMname,SStandard)
                    break

        for i in self.Standard: #获取材料的标准名称
            if i[1] == Standard:
                Standard_name = i[2]
            if i[1] == SStandard:
                SStandard_name = i[2]
        
        m_jsyq['standard']=Standard
        m_jsyq['name']=Mname
        for i in self.M_JSYQ:
            if (ID_Material in i[1]) and (eval(i[2]) or (i[2] in 'None      ')):
                m_jsyq['rcl']=i[0] #获取材料热处理状态
        return m_jsyq
   
if __name__ == '__main__': 
    tmjsyq=TMJSYQ()
    JSYQjiemianbianma=tmjsyq.get_JSYQmianban()
    print("JSYQjiemianbianma=\n",JSYQjiemianbianma)
    key=0
    for j in [tmjsyq.Material,tmjsyq.SMaterial,tmjsyq.Standard]:
        for i in j:
            if key%10==0:
                print(i)
            key+=1
'''
JSYQjiemianbianma的数据格式如下：
[2, 0, [1, 3], [2, 4], 4, [-20, 300], [4.0, -0.1], 3, 7, 2000, 30000, 20]
'''
