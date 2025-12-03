from openpyxl import load_workbook, Workbook

class Singleton:
    def __new__(cls,*args,**kwargs):
        if not hasattr(cls,"_instance"):
            cls._instance=super(Singleton,cls).__new__(cls)
        return cls._instance

class PartsData(Singleton):
    Kong = [None, "", " ", "  ", "   ", "     ", "      ", 'None',[],[None],['None']]
    """
    模块代码，模块名称英文，模块名称中文，可用父模块
    """
    key_mode = [
            [10, 'Nestle', '我的项目', [0]],
            [11, 'Equipment', '设备', [10]],
            [21, 'Component', '部件', [11, 21]],
            [51, 'Shell', '常用件', [11, 21]],
            [52, 'NozzleFlange', '管法兰', [11, 21]],
            [53, 'VesselFlange', '容器法兰', [11, 21]],
            [54, 'Pipe', '管子&管件', [11, 21]],
            [55, 'Support', '容器支座', [11, 21]],
            [56, 'ManWay', '人孔&手孔', [11, 21]],
            [57, 'LiftingLug', '吊耳', [11, 21]],
            [22, 'PlatFormAll', '塔器平台', [11]],
            [31, 'PlatFormVer', '单层塔平台', [22]],
            [32, 'PlatFormHor', '卧式平台', [11]],
            [23, 'TMJSYQ', '技术要求', [11]],
    ]
    equipment_list=[
            [1,	'常压容器', 'B-常压容器.dwg',   'B-常压容器.bmp',[1],(180,168)],
            [2, '压力容器', 'B-容器.dwg',       'B-容器.bmp',[2],(180,192)],
            [3, '换热器',   'B-换热器.dwg',     'B-换热器.bmp',[4,5],(180,232)],
            [4, '塔器',     'B-塔器.dwg',       'B-塔器.bmp',[7],(180,200)],
            [5, '夹套容器', 'B-夹套或盘管.dwg',   'B-夹套或盘管.bmp',[9,10],(180,208)],
            [6, '大型储罐', 'B-大型储罐.dwg',    'B-大型储罐.bmp',[8],(180,168)],
            [7, '球罐',	    'B-球罐.dwg',	'B-球罐.bmp',[6],(180,168)],
            [8, '疲劳容器', 'B-疲劳容器.dwg',    'B-疲劳容器.bmp',[3],(180,168)],
            ]

    tukuang_list=[
            ['A1','a1.DWG',((25,10),(831,584),(651,95))],
            ['A1V','a1v.DWG',((10,10),(584,816),(404,95))],
            ['A2x3','a2x3.DWG',((25,10),(1251,584),(1071,95))],
            ['A2x3V','a2x3v.DWG',((10,10),(584,1236),(404,95))],
            ]
    guankouzaihe=[
        ['a*2000*DN1','a*2000*DN1','a*1500*DN1','a*150*DN1*DN1','a*130*DN1*DN1','a*100*DN1*DN1'],
        ['a*80*DN2','a*80*DN2','a*60*DN2','a*0.25*DN2*DN2','a*0.20*DN2*DN2','a*0.16*DN2*DN2'],
        ]


    default_tpvs=[True, True, '120',True, True, '2.5', True, False, '', True, False, '',0.00,0.00]
    default_material=[True, False, True, False, '板材', 'GB/T 713-2014', 'Q345R', 7.85, True, False, True, False, '', '', 7.9]
    default_weight=['0', '0', 15, '1.15', '0', '0']
    default_platform_all=[['PlatFormAll', 1, '塔器平台', '塔器平台-->参数设置'],
                          ['8302', '5000','A1V', '1:10']
                          ]

    default_platform_ver=[['PlatFormVer',2, 'EL5000', '塔平台-->'],
                          ['2000＜OD≤3100', '30', '2632', True, '5000', '1000',
                           '1 级，2.0kN/m2', '悬臂式托架  XTJ', '', '', '210', '', '']
                          ]
    
    default_platform_hor=[['PlatFormHor', 3, '卧式设备平台', '卧式平台--> EL10000'],
                          ['8317', 'Ⅱ型（双侧）DN1000~3200', '10000', '8900', '7600', '7000', '2024',
                           '8000', '120', '900', '100', '200', '500', '8200', 'A1', '1:30']
                          ]
    default_data={
        101:[['Nestle',0,'我的项目','我的项目'],
             ['25006', '面向未来的大项目', '详细设计','工艺装置1（001）\n工艺装置2（002）\n工艺装置3（003）',
              '450','350', '-17.8','-28.6', '8(0.2g)','Ⅱ/第三组', 'C',]
            ],
        111:[['Equipment',1,'设备','设备 参数设置'],
            default_tpvs,default_material,
            ['冷凝器', 'E201', '工艺装置2（002）', '换热器', '工程图 A1', '1:20', '2001'],
            [], #给MXB预留
             #Image上只存数据，图像名称存取没有意义，因为图像上的控件位置数据仍然需要查取
            [None, 'Ⅱ/D2', None, '15', '1.0', '1.6', '--', '--', '68/128', '200', '86', '工艺气', '易爆、中度危害',
  '--', None, '--', '--', '--', None, None, '--', '--', None, None, None, None, None, '--', '--', None,
  None, '--', '--', '--', '--', '--', '--', '--', '--', 'TSG 21-2016《固定式压力容器安全技术监察规程》',
  'GB/T 151-2014 《热交换器》', 'GB/T 150.1~4-2011《压力容器》', 'NB/T 47015-2011《压力容器焊接规程》',
  'HG/T 20584-2020《钢制化工容器制造技术规范》',
  '--', None, None, '--', '--', '--', '--', '--', '--', '--', '--', '--', '--', '--', '--',
  None, None, None, None, None, None, None, None, None, None, '--', None, '--', '--', '--',
  '--', '--', '--', '--', '--', None, None, None, None, '1.1', '1.6', '--', '--', '158/108',
  '200', '132', '工艺气', '易爆、中度危害', '--', None, '--', '--', '--', None, None, '--',
  '--', None, None, None, None, None, None, None, None, None, None, '--', '--', None, None, None,
  None, None, None, None, None, None, None, None, None, None, None, None, None, '--', '--', '--',
  '--', '--', '--', '--', '--', '--', None, None, None, None, None, None, None, None, None, None,
  '--', None, '--', '--', '--', '--', None, None, None, None],
            [], #给Loacation预留
            default_weight,
            ],
        211:[['Component', 2, '组合件', '部件->组合件'],
             default_tpvs,default_material,
            ['组合件', '水/蒸汽', '800', '3.0', '1.0', True, False, False, False, False, False, True, True, False, False],
            ['', '组合件', '1', '--', '0.0', '0.0', '', 0.0, 0.0],
            [],  #Image
            [[False, True, '0.0, 0.0, 0.0', '0.0, 1.0, 0.0', '', '', True]],
            default_weight,
            ],
        511:[['Shell', 3, '筒体', '常用件->筒体'],
             default_tpvs,default_material,
              ['无', '筒体'],
              ['', '筒体 DN2000 t=16 H=6000', '1', 'Q345R', '4772.9', '4772.9', '', 4772.9, 0.0],
              ['2000', '0', '16', '0', '0', '0', '6000'],
              [[False, True, '0.0,40, 0', '0.0, 1.0, 0.0', '', '', True]]
             ],
        521:[['NozzleFlange', 4, '管法兰', '管法兰'],
             default_tpvs,default_material,
             ['HG/T 20615-2009', 'WN', '', 'PN50  CL300', 'RF', '100  4"', 'Sch.40  114.3×6.3'],
             ['HG/T 20615-2009', '法兰WN  PN50-DN100 RF Sch.40', '1', '16M II', '14.2', '14.2', '', 14.2, 0.0],
             ['114.3', '255', '200', '8-φ22配M20', '30.2', '101.7', '146', '84', '0', '0', '157.2', '2', '0', '0', '0', '0', '0', '0', '0', '0'],
             [[False, True, '0.0, 0.0, 0.0', '0.0, 1.0, 0.0', '', '', True]]
             ],
        571:[['LiftingLug', 5, '吊耳', '吊耳'], 
             [False, False, '', False, False, '', False, False, '', False, False, '', 0.0, 0.0], 
             [True, False, True, False, '板材', 'GB/T 713.2-2023', 'Q345R', 7.85, True, False, True, True, '垫板', 'S30408', 7.93], 
             ['HG/T 21574-2018', '顶部板式吊耳', 'TPB-4   10t', '16'], 
             ['HG/T 21574-2018', '吊耳TPB-4', '1', 'Q345R/S30408', '17.6', '17.6', '', 11.26, 6.34], 
             ['32', '100', '140', '105', '400', '200', '0', '0', '0', '0', '0', '0', '0', '0'], 
             [[False, True, '0.0, 0.0, 0.0', '0.0, 1.0, 0.0', '', '', True]]
             ],
       541:[['Pipe', 6, '管子&管件', '管子&管件'],
             default_tpvs,default_material,
             ['HG/T 20615-2009', 'WN', '', 'PN50  CL300', 'RF', '100  4"', 'Sch.40  114.3×6.3'],
             ['HG/T 20615-2009', '法兰WN  PN50-DN100 RF Sch.40', '1', '16M II', '14.2', '14.2', '', 14.2, 0.0],
             ['114.3', '255', '200', '8-φ22配M20', '30.2', '101.7', '146', '84', '0', '0', '157.2', '2', '0', '0', '0', '0', '0', '0', '0', '0'],
             [[False, True, '0.0, 0.0, 0.0', '0.0, 1.0, 0.0', '', '', True]]
             ],
        }
       

    """
    数据格式[直径，壁厚，复层壁厚，筒体长度，法兰标准，类型1，类型2（系列），
    PN，MFM，DN，接管标准，接管系列，Thk_Sch]
    """
    favorite_data=[0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    __Material = None
    __SMaterial = None
    __Standard = None
    __equipment_design_data = None
    __shell_data = None
    __PipeN = None
    __Pipe = None
    __PipeCTR = None
    __PipeMain = None
    __Tube = None
    __NozzleFlange = None
    __LiftingLugA = None
    __LiftingLugB = None
    pipe_type = None

    def __init__(self):
        pass

    @property
    def equipment_design_data(self):
        if self.__equipment_design_data==None:
            print("读取设备设计数据表数据库，请稍候...")
            wb = load_workbook('database/Main.xlsx',data_only=True)
            ws1 = wb["Equipment"]
            self.__equipment_design_data=[]
            NNr=79  #79
            NNc=11
            temp_data=[]
            for i in range(NNr):
                 b=[]
                 for j in range(NNc):
                      a=ws1.cell(row=i+5, column=j+9).value
                      if a in self.Kong:
                          a=''
                      else:
                          a=eval(a)
                      b.append(a)
                 temp_data.append(b)
            for i in self.equipment_list:
                temp2=[]
                for j in i[4]:
                    temp1=[]
                    for k in temp_data:
                        temp1.append(k[j])
                    temp2.append(temp1)
                self.__equipment_design_data.append([i,temp2])
        return self.__equipment_design_data

    @property
    def Material(self):
        if self.__Material==None:
            print("读取材料数据库，请稍候...")
            wb = load_workbook('database/Material.xlsx', data_only=True)
            ws1 = wb["Material"]
            self.__Material = []
            NNr = ws1.cell(row=1, column=2).value
            for i in range(NNr):
                a = [ws1.cell(row=i + 3, column=2).value,
                     ws1.cell(row=i + 3, column=3).value,
                     ws1.cell(row=i + 3, column=4).value,
                     ws1.cell(row=i + 3, column=5).value,
                     str(ws1.cell(row=i + 3, column=6).value),
                     ws1.cell(row=i + 3, column=7).value,
                     ws1.cell(row=i + 3, column=8).value,
                     ws1.cell(row=i + 3, column=9).value,
                     ws1.cell(row=i + 3, column=10).value,
                     ws1.cell(row=i + 3, column=11).value,
                     ws1.cell(row=i + 3, column=12).value,
                     ]
                self.__Material.append(a)
        return self.__Material

    @property
    def SMaterial(self):
        if self.__SMaterial==None:
            print("读取材料数据库，请稍候...")
            wb = load_workbook('database/Material.xlsx', data_only=True)
            ws2 = wb["SMaterial"]
            self.__SMaterial = []
            NNr = ws2.cell(row=1, column=2).value
            for i in range(NNr):
                a = [ws2.cell(row=i + 3, column=2).value,
                     ws2.cell(row=i + 3, column=3).value,
                     ws2.cell(row=i + 3, column=4).value,
                     ws2.cell(row=i + 3, column=5).value,
                     str(ws2.cell(row=i + 3, column=6).value),
                     ws2.cell(row=i + 3, column=7).value,
                     ws2.cell(row=i + 3, column=8).value
                     ]
                self.__SMaterial.append(a)
        return self.__SMaterial

    @property
    def Standard(self):
        if self.__Standard==None:
            print("读取标准数据库，请稍候...")
            wb = load_workbook('database/Material.xlsx', data_only=True)
            ws3 = wb["Standard"]     
            self.__Standard = []
            NNr = ws3.cell(row=1, column=2).value
            for i in range(NNr):
                a = [ws3.cell(row=i + 2, column=2).value,
                     ws3.cell(row=i + 2, column=3).value,
                     ws3.cell(row=i + 2, column=4).value,
                     ]
                self.__Standard.append(a)
        return self.__Standard

    @property
    def PipeN(self):
        if self.__PipeN==None:
            print("读取管子数据库，请稍候...")
            wb = load_workbook('database/Pipe.xlsx', data_only=True)
            ws1 = wb["PipeN"]     
            self.__PipeN = []
            NNr=ws1.cell(row=1, column=1).value
            NNc=ws1.cell(row=1, column=2).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws1.cell(row=i+3, column=j+1).value
                    b.append(str(a))
                self.__PipeN.append(b)
        return self.__PipeN

    @property
    def LiftingLugA(self):
        if self.__LiftingLugA==None:
            print("读取吊耳数据库，请稍候...")
            wb = load_workbook('database/Main.xlsx',data_only=True)
            ws1 = wb["LiftingLug"]
            self.__LiftingLugA=[]
            NNr=ws1.cell(row=1, column=2).value
            NNc=ws1.cell(row=1, column=3).value
            for i in range(NNr):
                 b=[]
                 for j in range(NNc):
                      a=ws1.cell(row=i+3, column=j+2).value
                      b.append(a)
                 self.__LiftingLugA.append(b)
        return self.__LiftingLugA
 
    @property
    def LiftingLugB(self):
        if self.__LiftingLugB==None:
            print("读取吊耳数据库，请稍候...")
            wb = load_workbook('database/LiftingLug.xlsx',data_only=True)
            ws1 = wb["LiftingLug"]
            self.__LiftingLugB=[]
            NNr=ws1.cell(row=1, column=2).value
            NNc=ws1.cell(row=1, column=3).value
            for i in range(NNr):
                 b=[]
                 for j in range(NNc):
                      a=ws1.cell(row=i+2, column=j+1).value
                      b.append(str(a))
                 if b[0] != "x":
                     self.__LiftingLugB.append(b[1:])
        return self.__LiftingLugB
    @property
    def PipeMain(self):
        if self.__PipeMain == None:
            print("读取管件面板数据库，请稍候...")
            wb = load_workbook('database/Main.xlsx',data_only=True)
            ws1 = wb["Pipe"]
            self.__PipeMain=[]
            NNr=ws1.cell(row=1, column=2).value
            NNc=ws1.cell(row=1, column=3).value
            for i in range(NNr):
                 b=[]
                 for j in range(NNc):
                      a=ws1.cell(row=i+3, column=j+3).value
                      b.append(str(a))
                 if b[0] != "x":
                     self.__PipeMain.append(b)
        return self.__PipeMain
    @property
    def Pipe(self):
#       print("self.__Pipe=",self.__Pipe)
        if self.__Pipe==None:
            print("读取管子数据库，请稍候...")
            wb = load_workbook('database/Pipe.xlsx', data_only=True)
            ws1 = wb["Pipe"]     
            self.__Pipe = []
            NNr=ws1.cell(row=1, column=1).value
            NNc=ws1.cell(row=1, column=2).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws1.cell(row=i+3, column=j+1).value
                    b.append(str(a))
                self.__Pipe.append(b)
        return self.__Pipe

    @property
    def Tube(self):
        if self.__Tube==None:
            print("创建换热管数据库，请稍候...")
            self.__Tube = []
            for i in [6,8,10,15,16,20,25,32,40,50,65,80,100,]:
                for j in self.PipeN:
                    if str(i) == str(j[1]) and j[0] == 'II系列':
                        Do=j[3]
                for j in self.Pipe:
                    if str(i) == str(j[2]) and j[1] == 'II系列':
                        self.__Tube.append([Do,j[4]])
        return self.__Tube
                    
    @property
    def PipeCTR(self):
#        print("self.__PipeCTR=",self.__PipeCTR)
        if self.__PipeCTR == None:
            print("读取管件数据库，请稍候...")
            wb = load_workbook('database/Pipe.xlsx', data_only=True)
            ws1 = wb["PipeC"]
            PipeC = []
            NNr=ws1.cell(row=1, column=1).value
            NNc=ws1.cell(row=1, column=2).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws1.cell(row=i+3, column=j+1).value
                    b.append(str(a))
                PipeC.append(b)
            pipe_type={}   #管件类型字典
            for i in range(34):
                a=ws1.cell(row=i+2, column=9).value
                b=ws1.cell(row=i+2, column=10).value
                if a not in self.Kong:
                    pipe_type[a]=b
            ws2 = wb["PipeT"] 
            PipeT = []
            NNr=ws2.cell(row=1, column=1).value
            NNc=ws2.cell(row=1, column=2).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws2.cell(row=i+4, column=j+1).value
                    b.append(str(a))
                PipeT.append(b)
            
            ws3 = wb["PipeR"] 
            PipeR = []
            NNr=ws2.cell(row=1, column=1).value
            NNc=ws2.cell(row=1, column=2).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws3.cell(row=i+4, column=j+1).value
                    b.append(str(a))
                PipeR.append(b)
            #将重叠数据进行展开
            self.__PipeCTR=[]
            for i in self.Pipe:
                self.__PipeCTR.append([pipe_type[11],'',i[1],i[2],'',i[3],'',''])  
                #数据格式[管件标准,类型,管子系列,公称直径1,公称直径2,壁厚,无缝/有缝,数据1,数据2...]
            for i in PipeC:
                for j in i[1].split('/'):
                    if i[4] not in self.Kong:
                       self.__PipeCTR.append([i[0],j,pipe_type[31],i[2],'','','',i[6]])  
                    if i[5] not in self.Kong:
                       self.__PipeCTR.append([i[0],j,pipe_type[33],i[2],'','','',i[6]]) 
            for i in PipeT:
                if i[2] not in self.Kong:
                    if i[4] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[52],pipe_type[31],i[0],'','',pipe_type[81],i[4]])  #长半径 90°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[52],pipe_type[31],i[0],'','',pipe_type[82],i[4]])
                    if i[5] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[51],pipe_type[31],i[0],'','',pipe_type[81],i[5]])  #长半径 45°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[51],pipe_type[31],i[0],'','',pipe_type[82],i[5]])
                    if i[6] not in self.Kong:   
                       self.__PipeCTR.append([pipe_type[14],pipe_type[53],pipe_type[31],i[0],'','',pipe_type[81],i[6]])  #长半径 180°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[53],pipe_type[31],i[0],'','',pipe_type[82],i[6]]) 
                    if i[7] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[54],pipe_type[31],i[0],'','',pipe_type[81],i[7]])  #短半径 90°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[54],pipe_type[31],i[0],'','',pipe_type[82],i[7]]) 
                    if i[8] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[55],pipe_type[31],i[0],'','',pipe_type[81],i[8]])  #短半径 180°弯头      
                       self.__PipeCTR.append([pipe_type[14],pipe_type[55],pipe_type[31],i[0],'','',pipe_type[82],i[8]])
                    if i[9] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[57],pipe_type[31],i[0],'','',pipe_type[81],i[9]])  #3D 90°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[57],pipe_type[31],i[0],'','',pipe_type[82],i[9]]) 
                    if i[10] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[56],pipe_type[31],i[0],'','',pipe_type[81],i[10]])  #3D 45°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[56],pipe_type[31],i[0],'','',pipe_type[82],i[10]]) 
                    if i[11] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[58],pipe_type[31],i[0],'','',pipe_type[81],i[11],i[12]])  #等径三通
                       self.__PipeCTR.append([pipe_type[14],pipe_type[58],pipe_type[31],i[0],'','',pipe_type[82],i[11],i[12]]) 
                       self.__PipeCTR.append([pipe_type[14],pipe_type[59],pipe_type[31],i[0],'','',pipe_type[81],i[11],i[12]])  #等径四通
                       self.__PipeCTR.append([pipe_type[14],pipe_type[59],pipe_type[31],i[0],'','',pipe_type[82],i[11],i[12]]) 
                    if i[13] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[60],pipe_type[31],i[0],'','',pipe_type[82],i[13]])  #管帽
                    if i[14] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[61],pipe_type[31],i[0],'','',pipe_type[82],i[14],i[15],i[16],i[18],i[19]])  #翻边短节(长型)
                       self.__PipeCTR.append([pipe_type[14],pipe_type[61],pipe_type[31],i[0],'','',pipe_type[82],i[14],i[15],i[17],i[18],i[19]])  #翻边短节(短型)
                if i[3] not in self.Kong:
                    if i[4] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[52],pipe_type[33],i[0],'','',pipe_type[81],i[4],i[4]])  #长半径 90°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[52],pipe_type[33],i[0],'','',pipe_type[82],i[4],i[4]])
                    if i[5] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[51],pipe_type[33],i[0],'','',pipe_type[81],i[5],i[5]])  #长半径 45°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[51],pipe_type[33],i[0],'','',pipe_type[82],i[5],i[5]])
                    if i[6] not in self.Kong:   
                       self.__PipeCTR.append([pipe_type[14],pipe_type[53],pipe_type[33],i[0],'','',pipe_type[81],i[6]])  #长半径 180°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[53],pipe_type[33],i[0],'','',pipe_type[82],i[6]]) 
                    if i[7] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[54],pipe_type[33],i[0],'','',pipe_type[81],i[7],i[7]])  #短半径 90°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[54],pipe_type[33],i[0],'','',pipe_type[82],i[7],i[7]]) 
                    if i[8] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[55],pipe_type[33],i[0],'','',pipe_type[81],i[8]])  #短半径 180°弯头      
                       self.__PipeCTR.append([pipe_type[14],pipe_type[55],pipe_type[33],i[0],'','',pipe_type[82],i[8]])
                    if i[9] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[57],pipe_type[33],i[0],'','',pipe_type[81],i[9],i[9]])  #3D 90°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[57],pipe_type[33],i[0],'','',pipe_type[82],i[9],i[9]]) 
                    if i[10] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[56],pipe_type[33],i[0],'','',pipe_type[81],i[10],i[10]])  #3D 45°弯头
                       self.__PipeCTR.append([pipe_type[14],pipe_type[56],pipe_type[33],i[0],'','',pipe_type[82],i[10],i[10]]) 
                    if i[11] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[58],pipe_type[33],i[0],'','',pipe_type[81],i[11],i[12]])  #等径三通
                       self.__PipeCTR.append([pipe_type[14],pipe_type[58],pipe_type[33],i[0],'','',pipe_type[82],i[11],i[12]]) 
                       self.__PipeCTR.append([pipe_type[14],pipe_type[59],pipe_type[33],i[0],'','',pipe_type[81],i[11],i[12]])  #等径四通
                       self.__PipeCTR.append([pipe_type[14],pipe_type[59],pipe_type[33],i[0],'','',pipe_type[82],i[11],i[12]]) 
                    if i[13] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[60],pipe_type[33],i[0],'','',pipe_type[82],i[13]])  #管帽
                    if i[14] not in self.Kong:
                       self.__PipeCTR.append([pipe_type[14],pipe_type[61],pipe_type[33],i[0],'','',pipe_type[82],i[14],i[15],i[16],i[18],i[19]])  #翻边短节(长型)
                       self.__PipeCTR.append([pipe_type[14],pipe_type[61],pipe_type[33],i[0],'','',pipe_type[82],i[14],i[15],i[17],i[18],i[19]])  #翻边短节(短型)
            for i in PipeR:
                if i[2] not in self.Kong:
                    if i[8] not in self.Kong:
                        self.__PipeCTR.append([pipe_type[14],pipe_type[63],pipe_type[31],i[0],i[4],'',pipe_type[81],i[8]])  #长半径异90°弯头
                        self.__PipeCTR.append([pipe_type[14],pipe_type[63],pipe_type[31],i[0],i[4],'',pipe_type[82],i[8]])
                    if i[9] not in self.Kong:
                        self.__PipeCTR.append([pipe_type[14],pipe_type[64],pipe_type[31],i[0],i[4],'',pipe_type[81],i[9],i[10]])  #异径三通
                        self.__PipeCTR.append([pipe_type[14],pipe_type[64],pipe_type[31],i[0],i[4],'',pipe_type[82],i[9],i[10]])
                        self.__PipeCTR.append([pipe_type[14],pipe_type[65],pipe_type[31],i[0],i[4],'',pipe_type[81],i[9],i[10]])  #异径四通
                        self.__PipeCTR.append([pipe_type[14],pipe_type[65],pipe_type[31],i[0],i[4],'',pipe_type[82],i[9],i[10]])
                    if i[11] not in self.Kong:
                        self.__PipeCTR.append([pipe_type[14],pipe_type[66],pipe_type[31],i[0],i[4],'',pipe_type[81],i[11]])  #同心异径管
                        self.__PipeCTR.append([pipe_type[14],pipe_type[66],pipe_type[31],i[0],i[4],'',pipe_type[82],i[11]])
                        self.__PipeCTR.append([pipe_type[14],pipe_type[67],pipe_type[31],i[0],i[4],'',pipe_type[81],i[11]])  #偏心异径管
                        self.__PipeCTR.append([pipe_type[14],pipe_type[67],pipe_type[31],i[0],i[4],'',pipe_type[82],i[11]])
                if i[4] not in self.Kong:
                    if i[8] not in self.Kong:
                        self.__PipeCTR.append([pipe_type[14],pipe_type[63],pipe_type[33],i[0],i[4],'',pipe_type[81],i[8]])  #长半径异90°弯头
                        self.__PipeCTR.append([pipe_type[14],pipe_type[63],pipe_type[33],i[0],i[4],'',pipe_type[82],i[8]])
                    if i[9] not in self.Kong:
                        self.__PipeCTR.append([pipe_type[14],pipe_type[64],pipe_type[33],i[0],i[4],'',pipe_type[81],i[9],i[10]])  #异径三通
                        self.__PipeCTR.append([pipe_type[14],pipe_type[64],pipe_type[33],i[0],i[4],'',pipe_type[82],i[9],i[10]])
                        self.__PipeCTR.append([pipe_type[14],pipe_type[65],pipe_type[33],i[0],i[4],'',pipe_type[81],i[9],i[10]])  #异径四通
                        self.__PipeCTR.append([pipe_type[14],pipe_type[65],pipe_type[33],i[0],i[4],'',pipe_type[82],i[9],i[10]])
                    if i[11] not in self.Kong:
                        self.__PipeCTR.append([pipe_type[14],pipe_type[66],pipe_type[33],i[0],i[4],'',pipe_type[81],i[11]])  #同心异径管
                        self.__PipeCTR.append([pipe_type[14],pipe_type[66],pipe_type[33],i[0],i[4],'',pipe_type[82],i[11]])
                        self.__PipeCTR.append([pipe_type[14],pipe_type[67],pipe_type[33],i[0],i[4],'',pipe_type[81],i[11]])  #偏心异径管
                        self.__PipeCTR.append([pipe_type[14],pipe_type[67],pipe_type[33],i[0],i[4],'',pipe_type[82],i[11]])
        #self.pipe_type = pipe_type
        return self.__PipeCTR

    @property
    def shell_data(self):
        if self.__shell_data==None:
            print("读取常用件数据库，请稍候...")
            wb = load_workbook('database/Main.xlsx',data_only=True)
            ws1 = wb["Shell"]
            self.__shell_data=[]
            NNr=ws1.cell(row=1, column=2).value
            NNc=ws1.cell(row=1, column=3).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws1.cell(row=i+3, column=j+2).value
                    b.append(a)
                self.__shell_data.append(b)
        return self.__shell_data

    @property
    def NozzleFlange(self):
        if self.__NozzleFlange==None:
#           print("读取管法兰数据库，请稍候...")
            wb = load_workbook('database/Main.xlsx',data_only=True)
            ws1 = wb["NozzleFlange"]
            flange=[]
            NNr=ws1.cell(row=1, column=2).value
            NNc=ws1.cell(row=1, column=3).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws1.cell(row=i+3, column=j+3).value
                    b.append(str(a))
                flange.append(b)
            self.__NozzleFlange=[]
            #将重叠数据进行展开
            for i in flange:
                if i[0] not in self.Kong:
                    for j in i[0].split(';'):
                        if i[2] not in self.Kong:
                            for k in i[2].split('/'):
                                t=[j,i[1],k,]
                                for m in i[3:]:
                                    t.append(m)
                                self.__NozzleFlange.append(t)
                        else:
                            t=[j,i[1],i[2],]
                            for m in i[3:]:
                                t.append(m)
                            self.__NozzleFlange.append(t)

            wb = load_workbook('database/NozzleFlange.xlsx',data_only=True)
            ws1 = wb["Class-PN"]
            ws2 = wb["RF"]
            ws3 = wb["MFMTG"]
            ws4 = wb["RJ"]
            ws5 = wb["Size"]
            ws6 = wb["All"]
#-----------------------------------------------------------
            self.NozzleFlange_class_pn=[]
            NNr=ws1.cell(row=1, column=1).value
            NNc=ws1.cell(row=1, column=2).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws1.cell(row=i+3, column=j+1).value
                    if a==None:
                        a=''
                    b.append(str(a))
                self.NozzleFlange_class_pn.append(b) 
#-----------------------------------------------------------
#           print('读取法兰密封面数据...')
            self.NozzleFlange_mfm=[]  #密封面数据
            NNr=ws2.cell(row=1, column=1).value
            NNc=ws2.cell(row=1, column=2).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws2.cell(row=i+3, column=j+1).value
                    if a==None:
                        a=''
                    b.append(str(a))
                if b[1] not in self.Kong:
                    for j in b[1].split('/'):
                        t=[b[0],j,'RF',]
                        for m in b[2:]:
                            t.append(m)
                        self.NozzleFlange_mfm.append(t)
            NNr=ws3.cell(row=1, column=1).value
            NNc=ws3.cell(row=1, column=2).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws3.cell(row=i+3, column=j+1).value
                    if a==None:
                        a=''
                    b.append(str(a))
                if b[1] not in self.NozzleFlange_mfm:
                    for j in b[1].split('/'):
                        for k in ['M','FM','T','G']:
                            t=[b[0],j,k,b[2],''] # 最后面的空格是补齐了Type数据列
                            for m in b[3:]:
                                t.append(m)
                            self.NozzleFlange_mfm.append(t)        
            for i in self.NozzleFlange_mfm:       #补充MFMTG中的d这个数据
                if i[5] in self.Kong:
                    for j in self.NozzleFlange_mfm:
                        if i[0]+i[1]+"RF"+i[3]==j[0]+j[1]+j[2]+j[3]:
                            i[5]=j[5]
                            break
            NNr=ws4.cell(row=1, column=1).value
            NNc=ws4.cell(row=1, column=2).value
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws4.cell(row=i+3, column=j+1).value
                    if a==None:
                        a=''
                    b.append(str(a))
                t=[b[0],b[1],'RJ',b[2],]
                for m in b[3:]:
                    t.append(m)
                self.NozzleFlange_mfm.append(t)
#-----------------------------------------------------------
#           print('读取法兰尺寸数据...')
            self.NozzleFlange_size=[]  #法兰尺寸数据
            NNr=ws5.cell(row=1, column=1).value
            NNc=ws5.cell(row=1, column=2).value
            flange_size=[]
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws5.cell(row=i+3, column=j+1).value
                    if a==None:
                        a=''
                    b.append(str(a))
                flange_size.append(b)
            NNr=ws6.cell(row=1, column=1).value
            NNc=ws6.cell(row=1, column=2).value
            flange_all=[]
            for i in range(NNr):
                b=[]
                for j in range(NNc):
                    a=ws6.cell(row=i+3, column=j+1).value
                    if a==None:
                        a=''
                    b.append(str(a))
                flange_all.append(b)
            for i in flange_all:
                for j in flange_size:
                    if i[0]+i[1]+i[2]+i[3]+i[5] == j[1]+j[2]+j[3]+j[4]+j[5]:
                        self.NozzleFlange_size.append(i+j[6:])
                        break
        return self.__NozzleFlange
 
    def Factory(self,key):
        '''
根据key返回对象，key可以是key_mode中的ID号，也可以是key_mode中的不区分大小写的名称。
替代a = eval(data[0][0] + "(self,data=data)")  这种构造nestle对象的办法
'''
        for i in self.key_mode:
            if key==i[0] or str(key).lower()== i[1].lower():
                return i[1]+"(self)" # 例如返回 "Equipment(self)"
                break
        return False

    def GetKeyModeIDList(self,key):
        '''根据key返nestle的ID和可用父nestle ID的列表
key可以是key_mode中的ID号，也可以是key_mode中的不区分大小写的名称。

'''
        for i in self.key_mode:
            if key==i[0] or str(key).lower()== i[1].lower():
                return (i[0],i[3])
                break
        return False
    
    def ToCombox(self, one_list):
        temp_list = []
        for one in one_list:
            if one not in temp_list:
                temp_list.append(one)
        return temp_list

    def bubbleSort(self, arr1):
        '''冒泡法排序'''
        arr = []
        for i in arr1:
            try:
                arr.append(int(i))
            except:
                pass
        for i in range(1, len(arr)):
            for j in range(0, len(arr) - i):
                if arr[j] > arr[j + 1]:
                    arr[j], arr[j + 1] = arr[j + 1], arr[j]
        return arr
    def dopipesheet(self):
        '''        输出管子数据到Excel表格        '''
        wb = Workbook()
        allpart_data=self  #PartsData()
        for i in ['Ia系列','Ib系列','II系列']:
            d1s=[['DN','Do']]
            key1=1
            das=['','']
            keya=1
            ws = wb.create_sheet(title=i)
            for hi in allpart_data.Pipe:
                if hi[1] == i:
                    d1 = [hi[2],hi[5]]
                    da = hi[3]
                    if d1 not in d1s:
                        d1s.append(d1)
                    if da not in das:
                        das.append(da)
                    drow = d1s.index(d1)+1
                    dcol = das.index(da)+1
                    ws.cell(row=drow, column=dcol).value=hi[4]
            for index,value in enumerate(das):
                ws.cell(row=1, column=index+1).value=value
            for index,value in enumerate(d1s):
                ws.cell(row=index+1, column=1).value=value[0]
                ws.cell(row=index+1, column=2).value=value[1]
        wb.save('database/接管壁厚表2.xlsx')
        wb.close()

if __name__ == '__main__': 
    allpart_data=PartsData()
    key=0
    i=0
    for hi in allpart_data.PipeCTR:
        i+=1
        j=0
        for hj in hi:
            j+=1
            #ws1.cell(row=i, column=j).value=hj
        if key%50==0:
            print(hi)
        key+=1
    for hi in allpart_data.PipeMain:
        print(hi)

