from openpyxl import load_workbook, Workbook
from ALLPartData import PartsData

class TMJSYQ():
    allpart_data=PartsData()
    list_equipment=[i[1] for i in allpart_data.equipment_list]
#list_equipment=['常压容器', '压力容器', '换热器', '塔器', '夹套容器', '大型储罐', '球罐', '疲劳容器']   
    list_type=["容器","壳程","管程","外夹套","外伴管","内伴管","内浮头","组合件",]
    list_texing=["易燃","易爆","中度危害","高度危害","极度危害"]
    list_fushi=["晶间腐蚀","氢腐蚀","湿H2S腐蚀","严重湿H2S腐蚀","NH3腐蚀"]
    list_leibie=["--","Ⅰ/D1","Ⅱ/D2","Ⅲ/A1","Ⅲ/A2","Ⅲ/A3","Ⅱ/A3",'Ⅰ/A3','Ⅱ/A1','Ⅰ/SAD','Ⅱ/SAD','Ⅲ/SAD']
    def __init__(self):
        print("读取材料技术要求数据库，请稍候...")
        wb = load_workbook('database/TMJSYQv5.xlsx', data_only=True)
        ws1 = wb["M_JSYQ"]
        self.M_JSYQ = []
        NNr=ws1.cell(row=9, column=2).value
        NNc=ws1.cell(row=9, column=3).value
        for i in range(NNr):
            b=[]
            for j in range(NNc):
                a=ws1.cell(row=i+11, column=j+2).value
                b.append(str(a))
            self.M_JSYQ.append(b)
        for i in self.M_JSYQ:
            i[1]=self.allpart_data.tools_page(i[1]) #将材料转化成列表
            
        self.Material = self.allpart_data.Material
        self.SMaterial = self.allpart_data.SMaterial
        self.Standard = self.allpart_data.Standard

    def get_material_jsyq(self,data):
        ID_Material = data.get('jiceng',None)
        ID_SMaterial =  data.get('fuceng',None)
        C =  data.get('C',None)
        Standard = None
        SStandard = None
        if ID_Material:
            for i in self.Material: #获取基层材料的类型和名称
                if i[0] == ID_Material:
                    Mtype =  i[1]  #材料类型
                    Standard = i[3]
                    Mname =  i[4]  #材料名称
                    Tem_dc_low = i[7]
                    Tem_dc_nom =  i[8]
                    Kv2 =  i[9]
        m_jsyq={}            
        if ID_SMaterial:
            for i in self.SMaterial: #获取复层材料的类型和名称
                if i[0] == ID_SMaterial:
                    SMtype =  i[1]  #复合层材料类型
                    SStandard = i[3]
                    SMname =  i[4]  #复合层材料名称
                    if '堆焊' in i[1]:
                        m_jsyq['fujia'] = '堆焊' + SMname
                    if '复合层' in i[1]:
                        m_jsyq['fujia'] = '复层材料{}应符合{}的规定'.format(SMname,SStandard)
                    break

        for i in self.Standard: #获取材料的标准名称
            if i[1] == Standard:
                Standard_name = i[2]
            if i[1] == SStandard:
                SStandard_name = i[2]
        
        m_jsyq['standard']=Standard
        m_jsyq['name']=Mname
        for i in self.M_JSYQ:
            if (ID_Material in i[1]) and (eval(i[2]) or (i[2] in 'None      ')):
                m_jsyq['rcl']=i[0] #获取材料热处理状态
        return m_jsyq

                
'''
        temstr="{}的{}所用{}{}应符合{}《{}》中的规定。"
        if ID_Material and Standard:
            temstr1= temstr.format(data.get('equipment'),
                data.get('type'),Mtype,Mname,Standard,Standard_name)
            print(temstr1)
            print('配套材料ID是：',Peitao)
        if ID_SMaterial and SStandard:
            temstr2= temstr.format(data.get('equipment'),
                data.get('type'),SMtype,SMname,SStandard,SStandard_name)
            print(temstr2) 
'''
# 受压元件 材料标准 材料牌号 热处理状态 附加要求

   
if __name__ == '__main__': 
    jsyq=TMJSYQ()  
#  a=[0, '板材', 59, 'GB/T 713-2014', 'Q245R', 7.85, 'CS', -20, 0, 34, '39,92']
    data={
        'equipment':"压力容器",
        'type1':"容器",
        'type2':"筒体",
        'texing':"易燃",
        'fushi':"晶间腐蚀",
        'leibie':"Ⅰ/D1",
        'jiceng':77,   #基层材料ID
        'fuceng':None, #复层材料ID
        'T':(200,),
        'P':(3.0,),
        'D':2000,
        'L':5000,
        'C':30,
        }
    print(jsyq.get_material_jsyq(data))
##    standardlist=[]
##    for i in jsyq.Material:
##        if i[3] not in standardlist:
##            standardlist.append(i[3])
##    for i in standardlist:
##        for j in jsyq.Standard:
##            if i == j[1]:
##                print( "{} 《{}》".format(i,j[2]))
##    
##    Namelist=[]
##    for i in jsyq.SMaterial:
##        if i[4] not in Namelist:
##            Namelist.append(i[4])
##    temlist=[]
##    for i in Namelist:
##        for  j in jsyq.Material:
##            if i == j[4]:
##                Mstandard=j[3]
##                Mname=j[4]
##                for k in jsyq.Standard:
##                    if k[1] == Mstandard:
##                        Mstandard_name=k[2]
##                        standardID=k[0]
##                temlist.append([standardID,Mname,Mstandard,Mstandard_name])
##    for i in temlist:
##        print( "{}  {}  {} 《{}》".format(i[0],i[1],i[2],i[3]))
##                    
                    
                

        
        
        
        
        
        
    
    
